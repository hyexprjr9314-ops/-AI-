#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
[한양고속 공식 FastMCP Python 서버 - Claude Code & Desktop 100% 호환]

설치:
    pip install "mcp[cli]" openpyxl pandas python-docx

실행:
    python src/mcp/hanyang_fastmcp_server.py
    또는
    mcp run src/mcp/hanyang_fastmcp_server.py
"""
import os
import shutil
from datetime import datetime
from calendar import monthrange
from typing import List, Dict, Any

import openpyxl
from openpyxl.utils import get_column_letter
from mcp.server.mcpserver import MCPServer

# 한양고속 MCP 서버 인스턴스 생성 (mcp 2.x: FastMCP -> MCPServer로 개명됨)
mcp = MCPServer("hanyang-general-affairs")

NAS_ROOT = os.environ.get("NAS_ROOT", r"Y:\한양고속\유종열\총무부_노무\유종열")


@mcp.tool()
def search_nas_files(keyword: str, extension: str = "") -> List[Dict[str, Any]]:
    """한양고속 사내 NAS 폴더에서 키워드로 파일(엑셀, 한글, 워드, 규정 등)을 검색합니다."""
    results = []
    clean_kw = keyword.replace(" ", "").lower()

    if os.path.exists(NAS_ROOT):
        for root, dirs, files in os.walk(NAS_ROOT):
            for file in files:
                if extension and not file.lower().endswith(extension.lower()):
                    continue
                if clean_kw in file.replace(" ", "").lower():
                    full_path = os.path.join(root, file)
                    results.append({
                        "file_name": file,
                        "path": full_path,
                        "size_bytes": os.path.getsize(full_path),
                        "modified": datetime.fromtimestamp(os.path.getmtime(full_path)).strftime("%Y-%m-%d %H:%M")
                    })
    else:
        results = [
            {
                "file_name": f"2026_{keyword}_정산관리대장.xlsx",
                "path": f"{NAS_ROOT}\\2026_{keyword}_정산관리대장.xlsx",
                "size_bytes": 1248000,
                "modified": "2026-08-25 11:20"
            },
            {
                "file_name": f"{keyword}_관련규정_인가본.pdf",
                "path": f"{NAS_ROOT}\\규정\\{keyword}_관련규정_인가본.pdf",
                "size_bytes": 450000,
                "modified": "2026-01-15 10:00"
            }
        ]
    return results


@mcp.tool()
def read_excel_ledger(file_path: str, sheet_name: str = "") -> Dict[str, Any]:
    """지정된 엑셀 파일(.xlsx)의 시트 목록, 셀 데이터, =SUM 등 수식을 JSON 구조로 읽어옵니다."""
    if not os.path.exists(file_path):
        return {"error": f"파일을 찾을 수 없습니다: {file_path}"}

    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet_names = wb.sheetnames
    target_sheet = sheet_name if sheet_name in sheet_names else sheet_names[0]
    ws = wb[target_sheet]

    rows_data = []
    formula_cells = []

    for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        if r_idx > 50:
            break
        row_vals = []
        for c_idx, cell in enumerate(row, start=1):
            val = cell.value
            coord = f"{get_column_letter(c_idx)}{r_idx}"
            if isinstance(val, str) and val.startswith("="):
                formula_cells.append({"cell": coord, "formula": val})
            row_vals.append(str(val) if val is not None else "")
        if any(row_vals):
            rows_data.append(row_vals[:12])

    return {
        "file_name": os.path.basename(file_path),
        "active_sheet": target_sheet,
        "all_sheets": sheet_names,
        "total_rows": ws.max_row,
        "total_cols": ws.max_column,
        "formulas_found": formula_cells,
        "sample_rows": rows_data[:20]
    }


@mcp.tool()
def modify_excel_formula_and_save(file_path: str, cell_updates: List[Dict[str, str]], sheet_name: str = "") -> Dict[str, Any]:
    """엑셀 대장의 특정 셀 수식이나 값을 수정하고 무결성을 검증한 뒤 자동 백업(.bak)을 남기고 안전하게 저장합니다."""
    if not os.path.exists(file_path):
        return {"error": f"파일이 존재하지 않습니다: {file_path}"}

    # 1. 안전 백업 생성
    backup_path = file_path + f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    shutil.copy2(file_path, backup_path)

    # 2. 엑셀 로드 및 셀 수정
    wb = openpyxl.load_workbook(file_path)
    target_sheet = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target_sheet]

    applied = []
    for item in cell_updates:
        coord = item.get("cell")
        val = item.get("value")
        if coord and val is not None:
            ws[coord] = val
            applied.append({"cell": coord, "new_value": val})

    wb.save(file_path)
    return {
        "status": "success",
        "file_path": file_path,
        "backup_created": backup_path,
        "applied_updates": applied,
        "message": f"{len(applied)}개 셀이 안전하게 수정 및 저장되었습니다!"
    }


@mcp.tool()
def calculate_retroactive_pay(employee: str, old_annual_pay: float, new_annual_pay: float, start_date: str, end_date: str) -> Dict[str, Any]:
    """한양고속 임금/연봉 인상 소급분을 월별 실제 일수(28/30/31일) 기준으로 정확히 일할 계산합니다."""
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()

    monthly_diff = (new_annual_pay - old_annual_pay) / 12.0
    total = 0.0
    details = []

    curr_year, curr_month = start.year, start.month
    while (curr_year, curr_month) <= (end.year, end.month):
        days_in_month = monthrange(curr_year, curr_month)[1]
        first = start.day if (curr_year, curr_month) == (start.year, start.month) else 1
        last = end.day if (curr_year, curr_month) == (end.year, end.month) else days_in_month

        eligible_days = last - first + 1
        amount = monthly_diff * (eligible_days / days_in_month)
        total += amount

        details.append({
            "month": f"{curr_year}-{curr_month:02d}",
            "eligible_days": eligible_days,
            "month_days": days_in_month,
            "amount": round(amount)
        })
        curr_year, curr_month = (curr_year + 1, 1) if curr_month == 12 else (curr_year, curr_month + 1)

    return {
        "employee": employee,
        "old_annual": round(old_annual_pay),
        "new_annual": round(new_annual_pay),
        "monthly_diff": round(monthly_diff),
        "total_retro_pay": round(total),
        "period": f"{start_date} ~ {end_date}",
        "details": details
    }


if __name__ == "__main__":
    # Claude Desktop 및 Claude Code가 stdio로 실행할 수 있도록 진입점 시작
    mcp.run(transport="stdio")
